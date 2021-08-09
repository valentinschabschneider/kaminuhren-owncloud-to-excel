from io import BytesIO
import logging
from pathlib import Path
import tkinter as tk
from tkinter import filedialog

from openpyxl.drawing.image import Image as XlImage
from openpyxl import Workbook
from openpyxl.styles import Alignment

from PIL import Image

FILE_TYPE_MAP = {"jpg": "jpeg"}


def load_image_thumbnail(path, format, size):
    """
    loads image into memory and compresses it
    """

    image = Image.open(path)
    image.thumbnail(size, Image.ANTIALIAS)

    memory_image = BytesIO()
    image.save(memory_image, format=format)
    return memory_image


def add_image_to_cell(cell, image_path: Path, size=(256, 256)):
    """
    loads image and pastes it into an excel cell
    """

    file_type = image_path.suffix[1:].lower()
    img = XlImage(
        load_image_thumbnail(
            image_path,
            FILE_TYPE_MAP[file_type] if file_type in FILE_TYPE_MAP else file_type,
            size,
        ),
    )
    worksheet.add_image(img, cell.coordinate)


# read root directory from user
root = tk.Tk()
root.withdraw()

path = filedialog.askdirectory(mustexist=True)

if not path:
    logging.info("No directory selected. Exiting.")
    exit()

root_directory = Path(path)

# only subfolders with naming scheme
clock_directories = [path for path in root_directory.glob("*-*") if path.is_dir()]

if not clock_directories:
    logging.warning('No matching folders in "%s". Exiting.', root_directory)
    exit()

# create Workbook
workbook = Workbook()
worksheet = workbook.worksheets[0]
worksheet.title = "Tabelle1"

for row_index, c_dir in enumerate(clock_directories, 1):
    # add folder name
    name_cell = worksheet.cell(row_index, 1)
    name_cell.value = c_dir.name

    # add description
    description_file_path = c_dir.joinpath("Beschreibung.txt")
    if not description_file_path.exists():
        logging.warning('"%s" does not exist!', description_file_path)
        description = ""
    else:
        try:
            description = open(description_file_path, encoding="utf-8").read()
        except UnicodeDecodeError:
            description = open(description_file_path, encoding="unicode_escape").read()

    description_cell = worksheet.cell(row_index, 2)
    description_cell.alignment = Alignment(wrapText=True)
    description_cell.value = description

    # add qr code image
    qr_code_file_path = c_dir.joinpath(c_dir.name).with_suffix(".png")
    if not qr_code_file_path.exists():
        logging.warning('"%s" does not exist!', qr_code_file_path)
    else:
        add_image_to_cell(worksheet.cell(row_index, 3), qr_code_file_path)

    # add clock image
    try:
        clock_image_file_path = next(c_dir.glob("*.jpg"))
    except StopIteration:
        logging.warning("no clock image found in %s", c_dir.name)
    else:
        add_image_to_cell(worksheet.cell(row_index, 4), clock_image_file_path)

    worksheet.row_dimensions[row_index].height = 192

# set column widths
worksheet.column_dimensions["A"].width = 14
worksheet.column_dimensions["B"].width = 70
worksheet.column_dimensions["C"].width = 36
worksheet.column_dimensions["D"].width = 36

# save file
file_path = Path(
    filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
        confirmoverwrite=True,
    )
)

workbook.save(file_path)
